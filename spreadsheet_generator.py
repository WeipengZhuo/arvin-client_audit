"""
Generate Excel spreadsheet deliverable with client assessments.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict
from datetime import datetime
from pathlib import Path


class SpreadsheetGenerator:
    """Generate formatted Excel spreadsheet with client assessments."""

    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Client Assessments"

        # Define styles
        self.header_font = Font(bold=True, size=12, color="FFFFFF")
        self.header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        self.cell_alignment = Alignment(vertical="top", wrap_text=True)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Color coding for recommendations
        self.color_continue = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
        self.color_cure = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
        self.color_terminate = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
        self.color_review = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # Blue

    def generate(self, analysis_results: List[Dict], output_path: str):
        """
        Generate formatted spreadsheet from analysis results.

        Args:
            analysis_results: List of client analysis dictionaries
            output_path: Path where spreadsheet will be saved
        """
        # Create headers
        self._create_headers()

        # Populate data rows
        self._populate_data(analysis_results)

        # Apply formatting
        self._apply_formatting()

        # Auto-adjust column widths
        self._adjust_column_widths()

        # Add summary sheet
        self._create_summary_sheet(analysis_results)

        # Save workbook
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(output_path)

        print(f"\n✅ Spreadsheet saved: {output_path}")

    def _create_headers(self):
        """Create column headers."""
        headers = [
            "Case Name",
            "PDF Source",
            "Client Classification",
            "Type of Notice Sent",
            "Firm Fault",
            "Firm Fault Explanation",
            "Current Status",
            "Recommendation",
            "Reasoning",
            "Key Evidence"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border

    def _populate_data(self, results: List[Dict]):
        """Populate data rows from analysis results."""
        for row_num, result in enumerate(results, 2):  # Start at row 2 (after headers)
            # Case Name
            self.ws.cell(row=row_num, column=1).value = result.get("case_name", "Unknown")

            # PDF Source
            self.ws.cell(row=row_num, column=2).value = result.get("pdf_filename", "Unknown")

            # Classification
            self.ws.cell(row=row_num, column=3).value = result.get("classification", "Not classified")

            # Notice Sent
            self.ws.cell(row=row_num, column=4).value = result.get("notice_sent", "Cannot determine")

            # Firm Fault
            firm_fault = result.get("firm_fault", "Unclear")
            self.ws.cell(row=row_num, column=5).value = firm_fault

            # Firm Fault Explanation
            self.ws.cell(row=row_num, column=6).value = result.get("firm_fault_explanation", "")

            # Current Status
            self.ws.cell(row=row_num, column=7).value = result.get("current_status", "Cannot determine")

            # Recommendation
            recommendation = result.get("recommendation", "Manual review required")
            cell = self.ws.cell(row=row_num, column=8)
            cell.value = recommendation

            # Apply color coding based on recommendation
            cell.fill = self._get_recommendation_color(recommendation)

            # Reasoning
            self.ws.cell(row=row_num, column=9).value = result.get("reasoning", "")

            # Key Evidence
            self.ws.cell(row=row_num, column=10).value = result.get("key_evidence", "")

    def _get_recommendation_color(self, recommendation: str) -> PatternFill:
        """Get color fill based on recommendation type."""
        rec_lower = recommendation.lower()

        if "continue" in rec_lower:
            return self.color_continue
        elif "cure" in rec_lower:
            return self.color_cure
        elif "terminat" in rec_lower:
            return self.color_terminate
        elif "review" in rec_lower or "executive" in rec_lower:
            return self.color_review
        else:
            return PatternFill()  # No fill

    def _apply_formatting(self):
        """Apply formatting to all data cells."""
        for row in self.ws.iter_rows(min_row=2, max_row=self.ws.max_row):
            for cell in row:
                cell.alignment = self.cell_alignment
                cell.border = self.border

                # Special formatting for Firm Fault column
                if cell.column == 5:  # Firm Fault column
                    if cell.value and "yes" in str(cell.value).lower():
                        cell.font = Font(bold=True, color="C00000")  # Bold red

    def _adjust_column_widths(self):
        """Auto-adjust column widths based on content."""
        column_widths = {
            1: 25,  # Case Name
            2: 30,  # PDF Source
            3: 20,  # Classification
            4: 25,  # Notice Sent
            5: 12,  # Firm Fault
            6: 40,  # Firm Fault Explanation
            7: 20,  # Current Status
            8: 30,  # Recommendation
            9: 50,  # Reasoning
            10: 50,  # Key Evidence
        }

        for col_num, width in column_widths.items():
            col_letter = get_column_letter(col_num)
            self.ws.column_dimensions[col_letter].width = width

        # Set row height for data rows
        for row in range(2, self.ws.max_row + 1):
            self.ws.row_dimensions[row].height = 60  # Taller rows for wrapped text

    def _create_summary_sheet(self, results: List[Dict]):
        """Create a summary sheet with statistics."""
        summary_ws = self.wb.create_sheet(title="Summary")

        # Title
        summary_ws["A1"] = "CLIENT ASSESSMENT SUMMARY"
        summary_ws["A1"].font = Font(bold=True, size=14)
        summary_ws.merge_cells("A1:B1")

        # Generation date
        summary_ws["A2"] = "Generated:"
        summary_ws["B2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Total cases
        summary_ws["A3"] = "Total Cases Analyzed:"
        summary_ws["B3"] = len(results)
        summary_ws["B3"].font = Font(bold=True)

        # Classification breakdown
        summary_ws["A5"] = "CLASSIFICATION BREAKDOWN"
        summary_ws["A5"].font = Font(bold=True)

        classifications = {}
        for result in results:
            classification = result.get("classification", "Unknown")
            classifications[classification] = classifications.get(classification, 0) + 1

        row = 6
        for classification, count in sorted(classifications.items()):
            summary_ws[f"A{row}"] = classification
            summary_ws[f"B{row}"] = count
            row += 1

        # Recommendation breakdown
        summary_ws[f"A{row + 1}"] = "RECOMMENDATION BREAKDOWN"
        summary_ws[f"A{row + 1}"].font = Font(bold=True)

        recommendations = {}
        for result in results:
            recommendation = result.get("recommendation", "Unknown")
            recommendations[recommendation] = recommendations.get(recommendation, 0) + 1

        row += 2
        for recommendation, count in sorted(recommendations.items()):
            cell_a = summary_ws[f"A{row}"]
            cell_b = summary_ws[f"B{row}"]

            cell_a.value = recommendation
            cell_b.value = count

            # Apply same color coding as main sheet
            cell_a.fill = self._get_recommendation_color(recommendation)

            row += 1

        # Firm fault breakdown
        summary_ws[f"A{row + 1}"] = "FIRM FAULT ANALYSIS"
        summary_ws[f"A{row + 1}"].font = Font(bold=True)

        firm_fault_yes = sum(1 for r in results if "yes" in str(r.get("firm_fault", "")).lower())
        firm_fault_no = sum(1 for r in results if "no" in str(r.get("firm_fault", "")).lower())
        firm_fault_unclear = len(results) - firm_fault_yes - firm_fault_no

        row += 2
        summary_ws[f"A{row}"] = "Firm Fault: YES"
        summary_ws[f"B{row}"] = firm_fault_yes
        summary_ws[f"B{row}"].font = Font(bold=True, color="C00000")

        summary_ws[f"A{row + 1}"] = "Firm Fault: NO"
        summary_ws[f"B{row + 1}"] = firm_fault_no

        summary_ws[f"A{row + 2}"] = "Firm Fault: UNCLEAR"
        summary_ws[f"B{row + 2}"] = firm_fault_unclear

        # Adjust column widths
        summary_ws.column_dimensions["A"].width = 35
        summary_ws.column_dimensions["B"].width = 15

        # Add legend
        row += 4
        summary_ws[f"A{row}"] = "COLOR LEGEND"
        summary_ws[f"A{row}"].font = Font(bold=True)

        row += 1
        legend_items = [
            ("Continue representation", self.color_continue),
            ("Send Notice to Cure", self.color_cure),
            ("Proceed with Termination", self.color_terminate),
            ("Executive Review Required", self.color_review),
        ]

        for label, color in legend_items:
            cell = summary_ws[f"A{row}"]
            cell.value = label
            cell.fill = color
            row += 1


def generate_spreadsheet(analysis_results: List[Dict], output_path: str):
    """
    Convenience function to generate spreadsheet.

    Args:
        analysis_results: List of analysis result dictionaries
        output_path: Where to save the Excel file
    """
    generator = SpreadsheetGenerator()
    generator.generate(analysis_results, output_path)
